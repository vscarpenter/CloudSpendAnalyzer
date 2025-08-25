"""Data export functionality for AWS Cost CLI."""

import csv
import json
import smtplib
from abc import ABC, abstractmethod
from datetime import datetime
from decimal import Decimal
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
from typing import Dict, Any, Optional, List, Union
import tempfile
import os

from .models import CostData, CostResult, CostAmount, QueryParameters


class DataExporter(ABC):
    """Abstract base class for data exporters."""
    
    @abstractmethod
    def export(self, cost_data: CostData, query_params: QueryParameters, output_path: str) -> str:
        """Export cost data to specified format."""
        pass


class CSVExporter(DataExporter):
    """CSV data exporter."""
    
    def export(self, cost_data: CostData, query_params: QueryParameters, output_path: str) -> str:
        """Export cost data to CSV format."""
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header with metadata
            writer.writerow(['# AWS Cost Data Export'])
            writer.writerow(['# Generated:', datetime.now().isoformat()])
            writer.writerow(['# Query:', getattr(query_params, 'original_query', 'N/A')])
            writer.writerow(['# Service:', query_params.service or 'All Services'])
            writer.writerow(['# Period:', f"{cost_data.time_period.start.date()} to {cost_data.time_period.end.date()}"])
            writer.writerow(['# Total Cost:', f"{cost_data.total_cost.amount} {cost_data.total_cost.unit}"])
            writer.writerow([])  # Empty row
            
            # Write main data headers
            headers = ['Period Start', 'Period End', 'Total Cost', 'Currency', 'Estimated']
            
            # Add group headers if available
            if cost_data.results and cost_data.results[0].groups:
                sample_group = cost_data.results[0].groups[0]
                if sample_group.keys:
                    headers.extend(['Group Keys', 'Group Cost'])
            
            writer.writerow(headers)
            
            # Write data rows
            for result in cost_data.results:
                base_row = [
                    result.time_period.start.date().isoformat(),
                    result.time_period.end.date().isoformat(),
                    float(result.total.amount),
                    result.total.unit,
                    result.estimated
                ]
                
                if result.groups:
                    # Write a row for each group
                    for group in result.groups:
                        row = base_row.copy()
                        if group.keys:
                            row.append(' / '.join(group.keys))
                            # Get primary cost metric
                            if group.metrics:
                                primary_cost = next(iter(group.metrics.values()))
                                row.append(float(primary_cost.amount))
                            else:
                                row.append(0.0)
                        writer.writerow(row)
                else:
                    writer.writerow(base_row)
            
            # Add trend analysis if available
            if cost_data.trend_data:
                writer.writerow([])  # Empty row
                writer.writerow(['# Trend Analysis'])
                writer.writerow(['Current Period Cost:', float(cost_data.trend_data.current_period.amount)])
                writer.writerow(['Comparison Period Cost:', float(cost_data.trend_data.comparison_period.amount)])
                writer.writerow(['Change Amount:', float(cost_data.trend_data.change_amount.amount)])
                writer.writerow(['Change Percentage:', cost_data.trend_data.change_percentage])
                writer.writerow(['Trend Direction:', cost_data.trend_data.trend_direction])
            
            # Add forecast data if available
            if cost_data.forecast_data:
                writer.writerow([])  # Empty row
                writer.writerow(['# Forecast Data'])
                writer.writerow(['Forecast Period Start', 'Forecast Period End', 'Forecasted Amount', 'Lower Bound', 'Upper Bound', 'Accuracy'])
                
                for forecast in cost_data.forecast_data:
                    writer.writerow([
                        forecast.forecast_period.start.date().isoformat(),
                        forecast.forecast_period.end.date().isoformat(),
                        float(forecast.forecasted_amount.amount),
                        float(forecast.confidence_interval_lower.amount),
                        float(forecast.confidence_interval_upper.amount),
                        forecast.prediction_accuracy or 'N/A'
                    ])
        
        return output_path


class JSONExporter(DataExporter):
    """JSON data exporter."""
    
    def export(self, cost_data: CostData, query_params: QueryParameters, output_path: str) -> str:
        """Export cost data to JSON format."""
        # Convert cost data to JSON-serializable format
        export_data = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'query': getattr(query_params, 'original_query', 'N/A'),
                'service': query_params.service,
                'granularity': query_params.granularity.value if hasattr(query_params.granularity, 'value') else str(query_params.granularity),
                'time_period': {
                    'start': cost_data.time_period.start.isoformat(),
                    'end': cost_data.time_period.end.isoformat()
                }
            },
            'summary': {
                'total_cost': {
                    'amount': float(cost_data.total_cost.amount),
                    'currency': cost_data.total_cost.unit
                },
                'currency': cost_data.currency,
                'group_definitions': cost_data.group_definitions
            },
            'results': []
        }
        
        # Add detailed results
        for result in cost_data.results:
            result_data = {
                'time_period': {
                    'start': result.time_period.start.isoformat(),
                    'end': result.time_period.end.isoformat()
                },
                'total': {
                    'amount': float(result.total.amount),
                    'currency': result.total.unit
                },
                'estimated': result.estimated,
                'groups': []
            }
            
            # Add group data
            for group in result.groups:
                group_data = {
                    'keys': group.keys,
                    'metrics': {}
                }
                for metric_name, cost_amount in group.metrics.items():
                    group_data['metrics'][metric_name] = {
                        'amount': float(cost_amount.amount),
                        'currency': cost_amount.unit
                    }
                result_data['groups'].append(group_data)
            
            export_data['results'].append(result_data)
        
        # Add trend analysis if available
        if cost_data.trend_data:
            export_data['trend_analysis'] = {
                'current_period': {
                    'amount': float(cost_data.trend_data.current_period.amount),
                    'currency': cost_data.trend_data.current_period.unit
                },
                'comparison_period': {
                    'amount': float(cost_data.trend_data.comparison_period.amount),
                    'currency': cost_data.trend_data.comparison_period.unit
                },
                'change': {
                    'amount': float(cost_data.trend_data.change_amount.amount),
                    'currency': cost_data.trend_data.change_amount.unit,
                    'percentage': cost_data.trend_data.change_percentage,
                    'direction': cost_data.trend_data.trend_direction
                }
            }
        
        # Add forecast data if available
        if cost_data.forecast_data:
            export_data['forecast'] = []
            for forecast in cost_data.forecast_data:
                forecast_data = {
                    'period': {
                        'start': forecast.forecast_period.start.isoformat(),
                        'end': forecast.forecast_period.end.isoformat()
                    },
                    'forecasted_amount': {
                        'amount': float(forecast.forecasted_amount.amount),
                        'currency': forecast.forecasted_amount.unit
                    },
                    'confidence_interval': {
                        'lower': {
                            'amount': float(forecast.confidence_interval_lower.amount),
                            'currency': forecast.confidence_interval_lower.unit
                        },
                        'upper': {
                            'amount': float(forecast.confidence_interval_upper.amount),
                            'currency': forecast.confidence_interval_upper.unit
                        }
                    },
                    'prediction_accuracy': forecast.prediction_accuracy
                }
                export_data['forecast'].append(forecast_data)
        
        # Write JSON file
        with open(output_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_data, jsonfile, indent=2, ensure_ascii=False)
        
        return output_path


class ExcelExporter(DataExporter):
    """Excel data exporter with charts and formatting."""
    
    def __init__(self):
        """Initialize Excel exporter."""
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Check if required dependencies are available."""
        try:
            import openpyxl
            self.openpyxl = openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    def export(self, cost_data: CostData, query_params: QueryParameters, output_path: str) -> str:
        """Export cost data to Excel format with charts and formatting."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, Reference
        from openpyxl.utils import get_column_letter
        
        # Create workbook and worksheets
        wb = self.openpyxl.Workbook()
        
        # Remove default sheet and create our sheets
        wb.remove(wb.active)
        summary_ws = wb.create_sheet("Summary")
        details_ws = wb.create_sheet("Detailed Data")
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        currency_format = '"$"#,##0.00'
        percentage_format = '0.00%'
        
        # Create Summary sheet
        self._create_summary_sheet(summary_ws, cost_data, query_params, header_font, header_fill, currency_format, percentage_format)
        
        # Create Detailed Data sheet
        self._create_details_sheet(details_ws, cost_data, header_font, header_fill, currency_format)
        
        # Add charts if we have time series data
        if len(cost_data.results) > 1:
            self._add_cost_chart(summary_ws, cost_data)
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, ws, cost_data, query_params, header_font, header_fill, currency_format, percentage_format):
        """Create the summary worksheet."""
        # Title and metadata
        ws['A1'] = "AWS Cost Data Export Summary"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A4'] = "Query:"
        ws['B4'] = getattr(query_params, 'original_query', 'N/A')
        
        ws['A5'] = "Service:"
        ws['B5'] = query_params.service or 'All Services'
        
        ws['A6'] = "Period:"
        ws['B6'] = f"{cost_data.time_period.start.date()} to {cost_data.time_period.end.date()}"
        
        # Total cost summary
        ws['A8'] = "Total Cost:"
        ws['B8'] = float(cost_data.total_cost.amount)
        ws['B8'].number_format = currency_format
        ws['B8'].font = Font(size=14, bold=True, color="008000")
        
        # Trend analysis if available
        if cost_data.trend_data:
            ws['A10'] = "Trend Analysis"
            ws['A10'].font = header_font
            ws['A10'].fill = header_fill
            
            ws['A11'] = "Current Period:"
            ws['B11'] = float(cost_data.trend_data.current_period.amount)
            ws['B11'].number_format = currency_format
            
            ws['A12'] = "Previous Period:"
            ws['B12'] = float(cost_data.trend_data.comparison_period.amount)
            ws['B12'].number_format = currency_format
            
            ws['A13'] = "Change Amount:"
            ws['B13'] = float(cost_data.trend_data.change_amount.amount)
            ws['B13'].number_format = currency_format
            
            ws['A14'] = "Change Percentage:"
            ws['B14'] = cost_data.trend_data.change_percentage / 100
            ws['B14'].number_format = percentage_format
            
            # Color code the change
            if cost_data.trend_data.change_percentage > 0:
                ws['B13'].font = Font(color="FF0000")  # Red for increase
                ws['B14'].font = Font(color="FF0000")
            elif cost_data.trend_data.change_percentage < 0:
                ws['B13'].font = Font(color="008000")  # Green for decrease
                ws['B14'].font = Font(color="008000")
        
        # Forecast summary if available
        if cost_data.forecast_data:
            start_row = 16 if cost_data.trend_data else 10
            
            ws[f'A{start_row}'] = "Cost Forecast"
            ws[f'A{start_row}'].font = header_font
            ws[f'A{start_row}'].fill = header_fill
            
            for i, forecast in enumerate(cost_data.forecast_data[:3], 1):
                row = start_row + i
                ws[f'A{row}'] = f"Month {i}:"
                ws[f'B{row}'] = float(forecast.forecasted_amount.amount)
                ws[f'B{row}'].number_format = currency_format
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_details_sheet(self, ws, cost_data, header_font, header_fill, currency_format):
        """Create the detailed data worksheet."""
        # Headers
        headers = ['Period Start', 'Period End', 'Total Cost', 'Currency', 'Estimated']
        
        # Add group headers if available
        if cost_data.results and cost_data.results[0].groups:
            headers.extend(['Service/Group', 'Group Cost'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = 2
        for result in cost_data.results:
            base_data = [
                result.time_period.start.date(),
                result.time_period.end.date(),
                float(result.total.amount),
                result.total.unit,
                'Yes' if result.estimated else 'No'
            ]
            
            if result.groups:
                # Write a row for each group
                for group in result.groups:
                    data = base_data.copy()
                    if group.keys:
                        data.append(' / '.join(group.keys))
                        # Get primary cost metric
                        if group.metrics:
                            primary_cost = next(iter(group.metrics.values()))
                            data.append(float(primary_cost.amount))
                        else:
                            data.append(0.0)
                    
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        if col == 3 or col == 7:  # Cost columns
                            cell.number_format = currency_format
                    row += 1
            else:
                for col, value in enumerate(base_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    if col == 3:  # Cost column
                        cell.number_format = currency_format
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_cost_chart(self, ws, cost_data):
        """Add a cost trend chart to the summary sheet."""
        try:
            from openpyxl.chart import LineChart, Reference
            
            # Create chart data in the worksheet
            chart_start_row = 20
            ws[f'A{chart_start_row}'] = "Period"
            ws[f'B{chart_start_row}'] = "Cost"
            
            for i, result in enumerate(cost_data.results, 1):
                row = chart_start_row + i
                ws[f'A{row}'] = result.time_period.start.strftime("%Y-%m-%d")
                ws[f'B{row}'] = float(result.total.amount)
            
            # Create chart
            chart = LineChart()
            chart.title = "Cost Trend"
            chart.style = 13
            chart.x_axis.title = "Time Period"
            chart.y_axis.title = "Cost (USD)"
            
            # Define data ranges
            data = Reference(ws, min_col=2, min_row=chart_start_row, max_row=chart_start_row + len(cost_data.results))
            categories = Reference(ws, min_col=1, min_row=chart_start_row + 1, max_row=chart_start_row + len(cost_data.results))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Add chart to worksheet
            ws.add_chart(chart, "D10")
            
        except Exception:
            # Skip chart creation if there's an error
            pass


class EmailReporter:
    """Email report functionality."""
    
    def __init__(self, smtp_config: Dict[str, Any]):
        """
        Initialize email reporter.
        
        Args:
            smtp_config: SMTP configuration dictionary with keys:
                - host: SMTP server host
                - port: SMTP server port
                - username: SMTP username
                - password: SMTP password
                - use_tls: Whether to use TLS (default: True)
        """
        self.smtp_config = smtp_config
    
    def send_report(self, cost_data: CostData, query_params: QueryParameters, 
                   recipients: List[str], subject: Optional[str] = None,
                   include_attachments: bool = True, attachment_formats: List[str] = None) -> bool:
        """
        Send cost report via email.
        
        Args:
            cost_data: Cost data to include in report
            query_params: Query parameters for context
            recipients: List of email addresses to send to
            subject: Email subject (auto-generated if None)
            include_attachments: Whether to include file attachments
            attachment_formats: List of formats to attach ('csv', 'json', 'excel')
        
        Returns:
            bool: True if email sent successfully
        """
        if attachment_formats is None:
            attachment_formats = ['csv', 'json']
        
        try:
            # Create message
            msg = MIMEMultipart()
            msg['From'] = self.smtp_config['username']
            msg['To'] = ', '.join(recipients)
            
            # Generate subject if not provided
            if subject is None:
                service_text = f" - {query_params.service}" if query_params.service else ""
                period_text = f"{cost_data.time_period.start.date()} to {cost_data.time_period.end.date()}"
                subject = f"AWS Cost Report{service_text} ({period_text})"
            
            msg['Subject'] = subject
            
            # Create email body
            body = self._create_email_body(cost_data, query_params)
            msg.attach(MIMEText(body, 'html'))
            
            # Add attachments if requested
            if include_attachments:
                attachments = self._create_attachments(cost_data, query_params, attachment_formats)
                for attachment_path, filename in attachments:
                    self._attach_file(msg, attachment_path, filename)
            
            # Send email
            with smtplib.SMTP(self.smtp_config['host'], self.smtp_config['port']) as server:
                if self.smtp_config.get('use_tls', True):
                    server.starttls()
                
                server.login(self.smtp_config['username'], self.smtp_config['password'])
                server.send_message(msg)
            
            # Clean up temporary files
            if include_attachments:
                for attachment_path, _ in attachments:
                    try:
                        os.unlink(attachment_path)
                    except:
                        pass
            
            return True
            
        except Exception as e:
            raise RuntimeError(f"Failed to send email report: {str(e)}")
    
    def _create_email_body(self, cost_data: CostData, query_params: QueryParameters) -> str:
        """Create HTML email body."""
        service_text = f" for {query_params.service}" if query_params.service else ""
        period_text = f"{cost_data.time_period.start.date()} to {cost_data.time_period.end.date()}"
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ background-color: #f0f0f0; padding: 15px; border-radius: 5px; }}
                .cost-summary {{ font-size: 24px; color: #2e7d32; font-weight: bold; margin: 20px 0; }}
                .trend {{ margin: 15px 0; }}
                .trend.up {{ color: #d32f2f; }}
                .trend.down {{ color: #2e7d32; }}
                .trend.stable {{ color: #f57c00; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .footer {{ margin-top: 30px; font-size: 12px; color: #666; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>AWS Cost Report{service_text}</h2>
                <p><strong>Period:</strong> {period_text}</p>
                <p><strong>Generated:</strong> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
            </div>
            
            <div class="cost-summary">
                Total Cost: ${cost_data.total_cost.amount:,.2f} {cost_data.total_cost.unit}
            </div>
        """
        
        # Add trend analysis if available
        if cost_data.trend_data:
            trend_class = cost_data.trend_data.trend_direction
            trend_symbol = "📈" if trend_class == "up" else "📉" if trend_class == "down" else "➡️"
            
            html += f"""
            <div class="trend {trend_class}">
                <h3>Trend Analysis {trend_symbol}</h3>
                <p><strong>Previous Period:</strong> ${cost_data.trend_data.comparison_period.amount:,.2f}</p>
                <p><strong>Change:</strong> ${cost_data.trend_data.change_amount.amount:,.2f} ({cost_data.trend_data.change_percentage:+.1f}%)</p>
            </div>
            """
        
        # Add detailed breakdown if available
        if len(cost_data.results) > 1:
            html += """
            <h3>Cost Breakdown by Period</h3>
            <table>
                <tr>
                    <th>Period</th>
                    <th>Cost</th>
                    <th>Status</th>
                </tr>
            """
            
            for result in cost_data.results:
                period = f"{result.time_period.start.date()} to {result.time_period.end.date()}"
                cost = f"${result.total.amount:,.2f}"
                status = "Estimated" if result.estimated else "Final"
                
                html += f"""
                <tr>
                    <td>{period}</td>
                    <td>{cost}</td>
                    <td>{status}</td>
                </tr>
                """
            
            html += "</table>"
        
        # Add service breakdown if available
        if cost_data.results and any(result.groups for result in cost_data.results):
            html += """
            <h3>Service Breakdown</h3>
            <table>
                <tr>
                    <th>Service/Resource</th>
                    <th>Cost</th>
                </tr>
            """
            
            # Get the result with the most groups
            best_result = max(cost_data.results, key=lambda r: len(r.groups) if r.groups else 0)
            
            # Sort groups by cost
            sorted_groups = sorted(
                best_result.groups,
                key=lambda g: max(cost.amount for cost in g.metrics.values()) if g.metrics else 0,
                reverse=True
            )
            
            for group in sorted_groups[:10]:  # Top 10 services
                if group.keys and group.metrics:
                    service_name = ' / '.join(group.keys)
                    primary_cost = next(iter(group.metrics.values()))
                    cost = f"${primary_cost.amount:,.2f}"
                    
                    html += f"""
                    <tr>
                        <td>{service_name}</td>
                        <td>{cost}</td>
                    </tr>
                    """
            
            html += "</table>"
        
        # Add forecast if available
        if cost_data.forecast_data:
            html += """
            <h3>Cost Forecast</h3>
            <table>
                <tr>
                    <th>Period</th>
                    <th>Forecasted Cost</th>
                    <th>Range</th>
                </tr>
            """
            
            for forecast in cost_data.forecast_data[:3]:
                period = f"{forecast.forecast_period.start.date()} to {forecast.forecast_period.end.date()}"
                forecasted = f"${forecast.forecasted_amount.amount:,.2f}"
                range_text = f"${forecast.confidence_interval_lower.amount:,.2f} - ${forecast.confidence_interval_upper.amount:,.2f}"
                
                html += f"""
                <tr>
                    <td>{period}</td>
                    <td>{forecasted}</td>
                    <td>{range_text}</td>
                </tr>
                """
            
            html += "</table>"
        
        html += """
            <div class="footer">
                <p>This report was generated by AWS Cost Explorer CLI.</p>
                <p>For questions or issues, please contact your system administrator.</p>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _create_attachments(self, cost_data: CostData, query_params: QueryParameters, 
                          formats: List[str]) -> List[tuple]:
        """Create temporary attachment files."""
        attachments = []
        
        with tempfile.TemporaryDirectory() as temp_dir:
            for format_type in formats:
                if format_type.lower() == 'csv':
                    exporter = CSVExporter()
                    filename = f"aws_cost_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    filepath = os.path.join(temp_dir, filename)
                    exporter.export(cost_data, query_params, filepath)
                    
                    # Copy to a permanent temp file
                    import shutil
                    perm_path = tempfile.mktemp(suffix='.csv')
                    shutil.copy2(filepath, perm_path)
                    attachments.append((perm_path, filename))
                
                elif format_type.lower() == 'json':
                    exporter = JSONExporter()
                    filename = f"aws_cost_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                    filepath = os.path.join(temp_dir, filename)
                    exporter.export(cost_data, query_params, filepath)
                    
                    # Copy to a permanent temp file
                    import shutil
                    perm_path = tempfile.mktemp(suffix='.json')
                    shutil.copy2(filepath, perm_path)
                    attachments.append((perm_path, filename))
                
                elif format_type.lower() == 'excel':
                    try:
                        exporter = ExcelExporter()
                        filename = f"aws_cost_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        filepath = os.path.join(temp_dir, filename)
                        exporter.export(cost_data, query_params, filepath)
                        
                        # Copy to a permanent temp file
                        import shutil
                        perm_path = tempfile.mktemp(suffix='.xlsx')
                        shutil.copy2(filepath, perm_path)
                        attachments.append((perm_path, filename))
                    except ImportError:
                        # Skip Excel if openpyxl is not available
                        pass
        
        return attachments
    
    def _attach_file(self, msg: MIMEMultipart, filepath: str, filename: str):
        """Attach a file to the email message."""
        with open(filepath, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename= {filename}'
        )
        
        msg.attach(part)


class ExportManager:
    """Main export manager that coordinates different exporters."""
    
    def __init__(self):
        """Initialize export manager."""
        self.exporters = {
            'csv': CSVExporter(),
            'json': JSONExporter(),
        }
        
        # Try to initialize Excel exporter
        try:
            self.exporters['excel'] = ExcelExporter()
        except ImportError:
            pass  # Excel export not available
    
    def export_data(self, cost_data: CostData, query_params: QueryParameters, 
                   format_type: str, output_path: str) -> str:
        """
        Export cost data to specified format.
        
        Args:
            cost_data: Cost data to export
            query_params: Query parameters for context
            format_type: Export format ('csv', 'json', 'excel')
            output_path: Output file path
        
        Returns:
            str: Path to exported file
        
        Raises:
            ValueError: If format is not supported
        """
        format_type = format_type.lower()
        
        if format_type not in self.exporters:
            available_formats = list(self.exporters.keys())
            raise ValueError(f"Unsupported export format '{format_type}'. Available formats: {available_formats}")
        
        exporter = self.exporters[format_type]
        return exporter.export(cost_data, query_params, output_path)
    
    def get_available_formats(self) -> List[str]:
        """Get list of available export formats."""
        return list(self.exporters.keys())
    
    def send_email_report(self, cost_data: CostData, query_params: QueryParameters,
                         smtp_config: Dict[str, Any], recipients: List[str],
                         subject: Optional[str] = None, attachment_formats: List[str] = None) -> bool:
        """
        Send email report with cost data.
        
        Args:
            cost_data: Cost data to include in report
            query_params: Query parameters for context
            smtp_config: SMTP configuration
            recipients: List of email addresses
            subject: Email subject (auto-generated if None)
            attachment_formats: List of formats to attach
        
        Returns:
            bool: True if email sent successfully
        """
        email_reporter = EmailReporter(smtp_config)
        return email_reporter.send_report(
            cost_data, query_params, recipients, subject, 
            include_attachments=bool(attachment_formats),
            attachment_formats=attachment_formats or []
        )